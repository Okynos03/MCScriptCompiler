import openpyxl
import xlrd
import tkinter as tk
from tkinter import filedialog


class Excel:
    def __init__(self):
        self.file_path = None
        self.root = tk.Tk()

    def open(self, path):
        self.file_path = path

    def read(self):
        list = []
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    row_values = [cell.value for cell in row]
                    if all(cell is None for cell in row_values):
                        break  # Stop reading after first completely empty row
                    list.append(row_values)

        except Exception as e:
            list = []
            try:
                workbook = xlrd.open_workbook(self.file_path)
                for sheet_name in workbook.sheet_names():
                    sheet = workbook.sheet_by_name(sheet_name)
                    for row_index in range(sheet.nrows):
                        row_values = sheet.row_values(row_index)
                        list.append(row_values)
            except Exception as e:
                print(f"Error reading the file: {e}")

        return list

    def read_asymmetrical(self):
        list = []
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    row_values = [cell.value for cell in row if cell.value is not None]
                    if all(cell is None for cell in row_values):
                        break
                    list.append(row_values)

        except Exception as e:
            list = []
            try:
                workbook = xlrd.open_workbook(self.file_path)
                for sheet_name in workbook.sheet_names():
                    sheet = workbook.sheet_by_name(sheet_name)
                    for row_index in range(sheet.nrows):
                        row_values = sheet.row_values(row_index)
                        list.append(row_values)
            except Exception as e:
                print(f"Error reading the file: {e}")

        return list


if __name__ == "__main__":
    excel = Excel()
    excel.open()
    print(excel.read())
